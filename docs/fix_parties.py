import openpyxl

BN_DIGITS = {'০':'0','১':'1','২':'2','৩':'3','৪':'4','৫':'5','৬':'6','৭':'7','৮':'8','৯':'9'}
def norm(s):
    return ''.join(BN_DIGITS.get(c,c) for c in str(s).strip())

wb_bn = openpyxl.load_workbook('/data/samples/mp_sample_bangla.xlsx')
ws_bn = wb_bn.active
wb_en = openpyxl.load_workbook('/data/samples/mp_sample_english.xlsx')
ws_en = wb_en.active

uid_to_bn = {}
for row in ws_bn.iter_rows(min_row=2, values_only=True):
    uid = norm(row[4])
    party = str(row[2] or '').strip()
    if uid and party:
        uid_to_bn[uid] = party

uid_to_en = {}
for row in ws_en.iter_rows(min_row=2, values_only=True):
    uid = norm(row[4])
    party = str(row[2] or '').strip()
    if uid and party:
        uid_to_en[uid] = party

mapping = {}
for uid, pbn in uid_to_bn.items():
    pen = uid_to_en.get(uid, '')
    if pen and pbn not in mapping:
        mapping[pbn] = pen

print(f'-- Party mapping: {len(mapping)} unique parties')
for bn, en in sorted(mapping.items(), key=lambda x: x[1]):
    bn_sql = bn.replace("'", "''")
    en_sql = en.replace("'", "''")
    print(f"UPDATE political_parties SET name_en = '{en_sql}' WHERE name_bn = '{bn_sql}';")
